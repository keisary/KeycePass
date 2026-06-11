"""
Script de génération des fichiers Excel de test pour KeycePass.
Produit un fichier .xlsx par classe à la racine du projet.

Format attendu par ImportService :
  Colonne 0 : matricule
  Colonne 1 : nom
  Colonne 2 : prenom
  Colonne 3 : classe_id
"""

import subprocess
import sys

# ── Installer openpyxl si absent ──────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("[*] openpyxl absent — installation en cours...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Etudiants de test : 2 classes ciblees (B2_IT & B2_MANAGEMENT) ─────────────
ETUDIANTS = [
    # B2_IT (10 etudiants)
    # matricule            nom              prenom           classe_id
    ("2026_B2IT_001",  "Koffi",         "Jean-Emmanuel",  "B2_IT"),
    ("2026_B2IT_002",  "Traore",        "Mariam",         "B2_IT"),
    ("2026_B2IT_003",  "Kouadio",       "Adjoua",         "B2_IT"),
    ("2026_B2IT_004",  "Diallo",        "Amadou",         "B2_IT"),
    ("2026_B2IT_005",  "Bamba",         "Kadiatou",       "B2_IT"),
    ("2026_B2IT_006",  "Coulibaly",     "Ibrahim",        "B2_IT"),
    ("2026_B2IT_007",  "Sissoko",       "Fatou",          "B2_IT"),
    ("2026_B2IT_008",  "Ndiaye",        "Aminata",        "B2_IT"),
    ("2026_B2IT_009",  "Toure",         "Mamadou",        "B2_IT"),
    ("2026_B2IT_010",  "Keita",         "Aicha",          "B2_IT"),

    # B2_MANAGEMENT (10 etudiants)
    ("2026_B2M_001",   "Ahou",          "Jean-Marc",      "B2_MANAGEMENT"),
    ("2026_B2M_002",   "Degny",         "Rachel",         "B2_MANAGEMENT"),
    ("2026_B2M_003",   "Ouattara",      "Grace",          "B2_MANAGEMENT"),
    ("2026_B2M_004",   "Soro",          "Herve",          "B2_MANAGEMENT"),
    ("2026_B2M_005",   "Kouakou",       "Irene",          "B2_MANAGEMENT"),
    ("2026_B2M_006",   "Kassi",         "Ketty",          "B2_MANAGEMENT"),
    ("2026_B2M_007",   "M'Boh",         "Leo",            "B2_MANAGEMENT"),
    ("2026_B2M_008",   "Adiko",         "Mireille",       "B2_MANAGEMENT"),
    ("2026_B2M_009",   "Boni",          "Noel",           "B2_MANAGEMENT"),
    ("2026_B2M_010",   "Yao",           "Sandra",         "B2_MANAGEMENT"),
]

# ── Couleurs ──────────────────────────────────────────────────────────────────
COULEUR_HEADER  = "1E3A5F"   # Bleu fonce
COULEUR_TEXTE_H = "FFFFFF"   # Blanc

COULEUR_CLASSE = {
    "B2_IT":         "D6EAF8",  # Bleu clair
    "B2_MANAGEMENT": "FDEDEC",  # Rouge clair
}

# ─────────────────────────────────────────────────────────────────────────────

def build_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def build_excel(output_path: str, etudiants: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Etudiants"

    headers       = ["matricule", "nom", "prenom", "classe_id"]
    header_font   = Font(bold=True, color=COULEUR_TEXTE_H, name="Calibri", size=11)
    header_fill   = PatternFill("solid", fgColor=COULEUR_HEADER)
    header_align  = Alignment(horizontal="center", vertical="center")
    border        = build_border()

    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border

    ws.row_dimensions[1].height = 20

    for row_idx, (matricule, nom, prenom, classe_id) in enumerate(etudiants, start=2):
        couleur_bg = COULEUR_CLASSE.get(classe_id, "FFFFFF")
        fill       = PatternFill("solid", fgColor=couleur_bg)

        for col_idx, val in enumerate([matricule, nom, prenom, classe_id], start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
            cell.font      = Font(name="Calibri", size=10)

        ws.row_dimensions[row_idx].height = 16

    for i, width in enumerate([20, 16, 18, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes       = "A2"
    ws.auto_filter.ref    = f"A1:D{len(etudiants) + 1}"
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    racine = os.path.dirname(os.path.abspath(__file__))

    # Grouper par classe (ordre de premiere apparition)
    classes: dict = {}
    for row in ETUDIANTS:
        classes.setdefault(row[3], []).append(row)

    # Un fichier par classe
    for classe_id, etudiants_classe in classes.items():
        nom_fichier = f"etudiants_{classe_id}.xlsx"
        sortie      = os.path.join(racine, nom_fichier)
        print(f"[*] Generation de {nom_fichier} ...")
        build_excel(sortie, etudiants_classe)
        noms = [f"{p} {n}" for _, n, p, _ in etudiants_classe]
        print(f"[OK] {nom_fichier} — {len(etudiants_classe)} etudiants : {', '.join(noms)}\n")

    print(f"[DONE] {len(classes)} fichier(s) genere(s) dans : {racine}")
